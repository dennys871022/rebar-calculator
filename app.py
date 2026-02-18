import streamlit as st
import pandas as pd
import math
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# === 1. 頁面設定 ===
st.set_page_config(page_title="鋼筋撿料大師 v23.0 (計算式完整版)", page_icon="🏗️", layout="wide")

# === 2. 初始化 Session State ===
if 'raw_data_list' not in st.session_state:
    st.session_state['raw_data_list'] = []

# === 3. 鋼筋基本資料 (CNS 560) ===
REBAR_DB = {
    '#3': {'dia': 0.953, 'weight': 0.560, 'db': 0.953},
    '#4': {'dia': 1.270, 'weight': 0.994, 'db': 1.27},
    '#5': {'dia': 1.590, 'weight': 1.560, 'db': 1.59},
    '#6': {'dia': 1.910, 'weight': 2.250, 'db': 1.91},
    '#7': {'dia': 2.220, 'weight': 3.040, 'db': 2.22},
    '#8': {'dia': 2.540, 'weight': 3.980, 'db': 2.54},
    '#9': {'dia': 2.870, 'weight': 5.060, 'db': 2.87},
    '#10': {'dia': 3.220, 'weight': 6.370, 'db': 3.22},
    '#11': {'dia': 3.580, 'weight': 7.907, 'db': 3.58}
}

# === 4. S7-01 標準圖數位化資料庫 ===
S7_DATA = {
    4200: {
        210: { 'tension_B': {'#3':55, '#4':73, '#5':91, '#6':110, '#7':158, '#8':181, '#9':205, '#10':229, '#11':255}, 'tension_Top': {'#3':71, '#4':95, '#5':118, '#6':142, '#7':206, '#8':235, '#9':266, '#10':298, '#11':331}, 'compression': {'#3':30, '#4':39, '#5':49, '#6':59, '#7':69, '#8':78, '#9':88, '#10':99, '#11':110}, 'develop': {'#3':42, '#4':56, '#5':70, '#6':84, '#7':122, '#8':139, '#9':157, '#10':177, '#11':196} },
        245: { 'tension_B': {'#3':51, '#4':68, '#5':85, '#6':101, '#7':147, '#8':168, '#9':189, '#10':212, '#11':236}, 'tension_Top': {'#3':66, '#4':88, '#5':110, '#6':132, '#7':190, '#8':218, '#9':246, '#10':276, '#11':307}, 'compression': {'#3':30, '#4':39, '#5':49, '#6':59, '#7':69, '#8':78, '#9':88, '#10':99, '#11':110}, 'develop': {'#3':39, '#4':52, '#5':65, '#6':78, '#7':113, '#8':129, '#9':146, '#10':164, '#11':182} },
        280: { 'tension_B': {'#3':48, '#4':63, '#5':79, '#6':95, '#7':137, '#8':157, '#9':177, '#10':199, '#11':221}, 'tension_Top': {'#3':62, '#4':82, '#5':103, '#6':123, '#7':178, '#8':204, '#9':230, '#10':258, '#11':287}, 'compression': {'#3':30, '#4':39, '#5':49, '#6':59, '#7':69, '#8':78, '#9':88, '#10':99, '#11':110}, 'develop': {'#3':37, '#4':49, '#5':61, '#6':73, '#7':106, '#8':121, '#9':136, '#10':153, '#11':170} },
        350: { 'tension_B': {'#3':43, '#4':57, '#5':71, '#6':85, '#7':123, '#8':140, '#9':159, '#10':178, '#11':198}, 'tension_Top': {'#3':55, '#4':74, '#5':92, '#6':110, '#7':159, '#8':182, '#9':206, '#10':231, '#11':257}, 'compression': {'#3':30, '#4':39, '#5':49, '#6':59, '#7':69, '#8':78, '#9':88, '#10':99, '#11':110}, 'develop': {'#3':33, '#4':44, '#5':55, '#6':65, '#7':95, '#8':108, '#9':122, '#10':137, '#11':152} }
    },
    2800: {
        210: { 'tension_B': {'#3':37, '#4':49, '#5':61, '#6':73, '#7':106, '#8':121, '#9':137, '#10':153, '#11':170}, 'tension_Top': {'#3':48, '#4':63, '#5':79, '#6':95, '#7':137, '#8':157, '#9':177, '#10':199, '#11':221}, 'compression': {'#3':30, '#4':30, '#5':33, '#6':40, '#7':46, '#8':52, '#9':59, '#10':66, '#11':74}, 'develop': {'#3':30, '#4':38, '#5':47, '#6':56, '#7':81, '#8':93, '#9':105, '#10':118, '#11':131} },
        280: { 'tension_B': {'#3':32, '#4':42, '#5':53, '#6':63, '#7':92, '#8':105, '#9':118, '#10':133, '#11':147}, 'tension_Top': {'#3':41, '#4':55, '#5':69, '#6':82, '#7':119, '#8':136, '#9':154, '#10':172, '#11':192}, 'compression': {'#3':30, '#4':30, '#5':33, '#6':40, '#7':46, '#8':52, '#9':59, '#10':66, '#11':74}, 'develop': {'#3':30, '#4':33, '#5':41, '#6':49, '#7':71, '#8':81, '#9':91, '#10':102, '#11':114} }
    }
}

# === 5. 核心查表功能 ===
def lookup_data(fc, fy, size, type_mode, is_top=False):
    val = 0; desc = "請手動輸入"
    try:
        fy_table = S7_DATA.get(fy)
        if fy_table:
            fc_table = fy_table.get(fc)
            if fc_table:
                key = ""
                if type_mode == 'compression': key = 'compression'
                elif type_mode == 'develop': key = 'develop'
                else: key = 'tension_Top' if is_top else 'tension_B'
                val = fc_table[key].get(size, 0)
                if val > 0: desc = "標準圖查表"
    except: pass
    if val == 0:
        db = REBAR_DB[size]['db']
        desc = "公式估算"
        if type_mode == 'compression': val = max(math.ceil(0.043 * fy * db), 20)
        else:
            factor = 46 * (fy / 4200) * math.sqrt(280 / fc)
            if is_top: factor *= 1.3
            val = math.ceil(factor * db * 1.3)
    return val, desc

# === 自動拆料演算法 ===
def split_rebar(req_len, stock_len, lap_len):
    if req_len <= stock_len:
        return [req_len]
    pieces = []
    rem = req_len
    while rem > stock_len:
        pieces.append(stock_len)
        rem -= (stock_len - lap_len)
    pieces.append(rem)
    return pieces

# === 刪除與清空 ===
def delete_item(raw_idx): 
    st.session_state['raw_data_list'].pop(raw_idx)
def clear_all_data(): 
    st.session_state['raw_data_list'] = []

# === 繪圖 ===
def plot_section(shape, dims, cover):
    fig, ax = plt.subplots(figsize=(3, 3))
    if shape == 'rect':
        w, h = dims['w'], dims['h']
        rect_conc = patches.Rectangle((0, 0), w, h, linewidth=2, edgecolor='#333333', facecolor='#f0f0f0')
        ax.add_patch(rect_conc)
        if w > 2*cover and h > 2*cover:
            rect_stir = patches.Rectangle((cover, cover), w-2*cover, h-2*cover, linewidth=1.5, edgecolor='red', linestyle='--', facecolor='none')
            ax.add_patch(rect_stir)
        ax.set_xlim(-10, w+10); ax.set_ylim(-10, h+10)
    elif shape == 'circle':
        d = dims['d']; r = d / 2
        circ_conc = patches.Circle((r, r), r, linewidth=2, edgecolor='#333333', facecolor='#f0f0f0')
        ax.add_patch(circ_conc)
        if r > cover:
            circ_stir = patches.Circle((r, r), r-cover, linewidth=1.5, edgecolor='red', linestyle='--', facecolor='none')
            ax.add_patch(circ_stir)
        ax.set_xlim(-10, d+10); ax.set_ylim(-10, d+10)
    ax.set_aspect('equal'); ax.axis('off')
    return fig

# === 8. 側邊欄設定 ===
with st.sidebar:
    st.header("⚙️ 專案參數設定")
    project_name = st.text_input("建案名稱", value="CDC防疫中心")
    contact_person = st.text_input("聯絡人", value="范嘉文")
    structure_part = st.text_input("結構部位", value="洗車台")
    st.markdown("---")
    
    st.subheader("2. 材料強度")
    fc_mode = st.selectbox("混凝土 f'c", [210, 245, 280, 350, "自訂"], index=2)
    fc = st.number_input("輸入 f'c", value=280, step=5) if fc_mode == "自訂" else fc_mode
    fy_mode = st.selectbox("鋼筋 f_y", [2800, 4200, "自訂"], index=1)
    fy = st.number_input("輸入 f_y", value=4200, step=100) if fy_mode == "自訂" else fy_mode
    
    stock_len = st.selectbox("鋼筋定尺 (m)", [9, 10, 12, 14, 15], index=2) * 100
    global_cover = st.number_input("預設保護層 (cm)", value=4.0, step=0.5)
    unit_price = st.number_input("鋼筋單價 (元/噸)", value=23000, step=500)
    
    st.markdown("---")
    st.subheader("3. 動態撿料設定")
    auto_split = st.checkbox("✅ 啟用自動拆料", value=True, help="變更此選項，右側表格會瞬間切換顯示模式！")

# === 9. 主畫面 ===
st.title("🏗️ 鋼筋撿料大師 v23.0")
st.caption(f"即時互動引擎: {'🚀 啟用自動拆料' if auto_split else '📦 關閉拆料(合併顯示)'} | 定尺 {stock_len/100}m")

with st.expander("➕ 新增撿料項目", expanded=True):
    col_input, col_viz = st.columns([2, 1])
    
    with col_input:
        c1, c2, c3 = st.columns([1.5, 1, 1])
        with c1: note_input = st.text_input("備註/結構部位", value=structure_part) 
        with c2: size_key = st.selectbox("番號", list(REBAR_DB.keys()), index=3)
        with c3: cover = st.number_input("保護層", value=global_cover)
        
        mode = st.radio("模式選擇", ["主筋 (梁/柱直料)", "版/牆筋 (依間距)", "箍筋 (Stirrup)", "螺旋箍筋 (Spiral)"], horizontal=True)
        
        db = REBAR_DB[size_key]['db']
        h90 = math.ceil(max(12*db, 15)); h180 = math.ceil(max(4*db, 6.5))
        
        suggested_lap = 0; lap_desc = ""; is_top = False
        if "主筋" in mode or "版/牆" in mode:
            is_col = st.checkbox("是柱子/受壓構件?", value=False)
            if not is_col:
                is_top = st.checkbox("頂層筋 (Top Bar)?", value=False)
                suggested_lap, lap_desc = lookup_data(fc, fy, size_key, 'tension', is_top)
            else:
                suggested_lap, lap_desc = lookup_data(fc, fy, size_key, 'compression')
        
        inputs = {}
        
        if "螺旋" in mode: 
            st.info("🌀 螺旋箍筋 (搭接 1.5 圈)")
            c_a, c_b = st.columns(2)
            with c_a: inputs['D'] = st.number_input("圓柱直徑 D (cm)", min_value=0.0)
            with c_b: inputs['L'] = st.number_input("樁長 L (cm)", min_value=0.0)
            c1, c2 = st.columns(2)
            with c1: inputs['P'] = st.number_input("間距 Pitch (cm)", value=15.0)
            with c2: inputs['count'] = st.number_input("總支數", min_value=1, value=1)
            
            if inputs['D'] > 0 and inputs['P'] > 0:
                core_d = inputs['D'] - 2*cover
                circ = math.pi * core_d
                one_turn = math.sqrt(circ**2 + inputs['P']**2)
                suggested_lap_spiral = 1.5 * one_turn
            else: suggested_lap_spiral = 0
            
            st.markdown(f"👇 **搭接設定 (建議值: 1.5圈)**")
            inputs['manual_lap'] = st.number_input("搭接長度", value=float(f"{suggested_lap_spiral:.1f}"), step=1.0)

        elif "主筋" in mode:
            c_a, c_b = st.columns(2)
            with c_a: inputs['L'] = st.number_input("實際跨距/淨長 (cm)", min_value=0.0)
            with c_b: inputs['count'] = st.number_input("支數", min_value=1, value=1)
            st.markdown(f"👇 **搭接設定 ({lap_desc})**")
            inputs['manual_lap'] = st.number_input("搭接長度", value=int(suggested_lap), step=1, key=f"lap_main_{fc}_{fy}_{size_key}_{is_top}_{is_col}")
            c_c, c_d = st.columns(2)
            with c_c: inputs['hL'] = st.selectbox("左鉤", ["平切", "90度", "180度"])
            with c_d: inputs['hR'] = st.selectbox("右鉤", ["平切", "90度", "180度"])

        elif "版/牆" in mode:
            c_a, c_b = st.columns(2)
            with c_a: inputs['L'] = st.number_input("實際跨距/淨長 (cm)", min_value=0.0)
            c_range, c_space = st.columns(2)
            with c_range: inputs['range_len'] = st.number_input("佈筋範圍 (cm)", min_value=0.0)
            with c_space: inputs['spacing'] = st.number_input("間距 @ (cm)", min_value=1.0, value=15.0)
            
            r_len = inputs['range_len']
            spc = inputs['spacing']
            calc_count = math.ceil(r_len / spc) + 1 if r_len > 0 else 1
            inputs['count'] = st.number_input("總支數 (自動計算)", value=int(calc_count), min_value=1, key=f"count_slab_{r_len}_{spc}")
            st.markdown(f"👇 **搭接設定 ({lap_desc})**")
            inputs['manual_lap'] = st.number_input("搭接長度", value=int(suggested_lap), step=1, key=f"lap_slab_{fc}_{fy}_{size_key}_{is_top}")
            c_c, c_d = st.columns(2)
            with c_c: inputs['hL'] = st.selectbox("左鉤", ["平切", "90度", "180度"])
            with c_d: inputs['hR'] = st.selectbox("右鉤", ["平切", "90度", "180度"])

        elif "箍筋" in mode:
            c_a, c_b = st.columns(2)
            with c_a: inputs['W'] = st.number_input("寬 W", min_value=0.0)
            with c_b: inputs['H'] = st.number_input("深 H", min_value=0.0)
            st_mode = st.radio("計算", ["智慧分區", "手動輸入"])
            if st_mode == "智慧分區":
                inputs['Span'] = st.number_input("淨跨距 L", min_value=0.0)
                c1, c2 = st.columns(2)
                with c1: inputs['sE'] = st.number_input("加密區 @ (填0表無加密)", value=10.0, min_value=0.0)
                with c2: inputs['sC'] = st.number_input("一般區 @", value=15.0, min_value=1.0)
                inputs['st_type'] = 'auto'
            else:
                inputs['count'] = st.number_input("總支數", min_value=1)
                inputs['st_type'] = 'manual'

        btn_add = st.button("➕ 加入清單", type="primary", use_container_width=True)

    # === 右側：視覺化與動態計算式預覽 ===
    with col_viz:
        st.markdown("#### 📐 計算式預覽")
        if "螺旋" in mode:
            d_val = inputs.get('D', 0)
            if d_val > 0:
                st.pyplot(plot_section('circle', {'d': d_val}, cover))
                core_d = d_val - 2*cover
                circ = math.pi * core_d
                p_val = inputs.get('P', 15.0)
                one_turn = math.sqrt(circ**2 + p_val**2) if p_val > 0 else 0
                st.latex(rf"D_{{core}} = {d_val} - 2({cover}) = {core_d} \text{{ cm}}")
                st.latex(rf"L_{{turn}} = \sqrt{{(\pi \times {core_d:.1f})^2 + {p_val}^2}} = {one_turn:.1f} \text{{ cm}}")
                if one_turn > 0:
                    st.latex(rf"L_{{lap}} (1.5\text{{圈}}) = 1.5 \times {one_turn:.1f} = \mathbf{{{1.5*one_turn:.1f} cm}}")
                
        elif "箍筋" in mode:
            w_val = inputs.get('W', 0)
            h_val = inputs.get('H', 0)
            if w_val > 0 and h_val > 0:
                st.pyplot(plot_section('rect', {'w': w_val, 'h': h_val}, cover))
                cw = w_val - 2*cover
                ch = h_val - 2*cover
                hook_s = max(24*db, 20)
                L_stirrup = (cw+ch)*2 + hook_s
                st.latex(rf"L_{{core}} = 2 \times ({cw} + {ch}) = {(cw+ch)*2} \text{{ cm}}")
                st.latex(rf"L_{{hook}} (135^\circ) = \max(24d_b, 20) = {hook_s} \text{{ cm}}")
                st.latex(rf"L_{{1\text{{支}}}} = {(cw+ch)*2} + {hook_s} = {L_stirrup} \text{{ cm}}")
                
                if inputs.get('st_type') == 'auto' and inputs.get('Span', 0) > 0:
                    span = inputs.get('Span', 0)
                    sE = inputs.get('sE', 10.0)
                    sC = inputs.get('sC', 15.0)
                    st.markdown("**支數分配：**")
                    if sE <= 0:
                        st.latex(rf"N_{{total}} = \lceil {span} / {sC} \rceil + 1")
                    else:
                        zE = 2 * h_val
                        if zE * 2 >= span:
                            st.latex(rf"Z_E (2H) \times 2 \ge L \rightarrow \text{{全加密}}")
                            st.latex(rf"N_{{total}} = \lceil {span} / {sE} \rceil + 1")
                        else:
                            zC_len = span - 2*zE
                            st.latex(rf"Z_{{\text{{加密}}}} = 2 \times H = {zE} \text{{ cm}}")
                            st.latex(rf"N_{{\text{{端}}}} = 2 \times \lceil {zE} / {sE} \rceil")
                            st.latex(rf"N_{{\text{{中}}}} = \lceil {zC_len} / {sC} \rceil")

        elif "主筋" in mode or "版/牆" in mode:
            if "版/牆" in mode and inputs.get('count', 0) > 0:
                r_len = inputs.get('range_len', 0)
                spc = inputs.get('spacing', 15.0)
                st.latex(rf"N_{{\text{{支數}}}} = \lceil {r_len} / {spc} \rceil + 1 = {inputs['count']}")

            l_val = inputs.get('L', 0)
            if l_val > 0:
                net = l_val - 2*cover
                hl_str = inputs.get('hL', '平切')
                hr_str = inputs.get('hR', '平切')
                hook_l = h90 if hl_str=="90度" else (h180 if hl_str=="180度" else 0)
                hook_r = h90 if hr_str=="90度" else (h180 if hr_str=="180度" else 0)
                req_len = net + hook_l + hook_r
                
                st.latex(rf"L_{{net}} = {l_val} - 2({cover}) = {net} \text{{ cm}}")
                if hook_l > 0 or hook_r > 0: 
                    st.latex(rf"L_{{hook}} = {hook_l} + {hook_r} = {hook_l + hook_r} \text{{ cm}}")
                st.latex(rf"L_{{req}} = {req_len} \text{{ cm (物理展開長)}}")
                
                if req_len > stock_len:
                    st.error(f"⚠️ 長度超過定尺 ({stock_len/100}m)！", icon="✂️")
                    if auto_split:
                        st.caption("啟動自動拆料：將自動切割為定尺與餘料搭接段。")
                    else:
                        st.caption("未啟動拆料：將把所有搭接長度合併顯示於同一筆。")

    # === 運算加入邏輯 (儲存原始資料) ===
    if btn_add:
        try:
            uw = REBAR_DB[size_key]['weight']
            hook_map = {"平切": 0, "90度": h90, "180度": h180}
            base_len = 0
            final_count = inputs.get('count', 1)
            shape_str = ""
            sys_mode = "main"

            if "主筋" in mode or "版/牆" in mode:
                if inputs['L'] <= 0: raise ValueError("長度需大於0")
                net = inputs['L'] - (2 * cover)
                base_len = net + hook_map[inputs.get('hL', '平切')] + hook_map[inputs.get('hR', '平切')]
                shape_str = f"L={inputs['L']}"
                sys_mode = "main"

            elif "螺旋" in mode:
                if inputs['D'] <= 0 or inputs['P'] <= 0: raise ValueError("請輸入正確尺寸")
                core_d = inputs['D'] - 2*cover
                circ = math.pi * core_d
                one_turn = math.sqrt(circ**2 + inputs['P']**2)
                base_len = (one_turn * (inputs['L'] / inputs['P'])) + (3.0 * circ)
                shape_str = f"◎ D={inputs['D']}"
                sys_mode = "main"

            elif "箍筋" in mode:
                cw = inputs['W'] - 2*cover; ch = inputs['H'] - 2*cover
                base_len = (cw+ch)*2 + max(24*db, 20)
                if inputs['st_type'] == 'auto':
                    if inputs['sE'] <= 0: final_count = math.ceil(inputs['Span'] / inputs['sC']) + 1
                    else:
                        zE = 2*inputs['H']
                        if zE*2 >= inputs['Span']: final_count = math.ceil(inputs['Span']/inputs['sE']) + 1
                        else:
                            final_count = math.ceil(zE/inputs['sE'])*2 + math.ceil((inputs['Span'] - 2*zE)/inputs['sC']) + 1
                shape_str = f"口 {inputs['W']}x{inputs['H']}"
                sys_mode = "stirrup"

            st.session_state['raw_data_list'].append({
                "mode": sys_mode, "size_key": size_key, "shape_str": shape_str,
                "base_len": base_len, "lap_len": inputs.get('manual_lap', 0),
                "count": final_count, "uw": uw, "note": note_input
            })
            st.success("已加入原始資料！請見下方報表。")
            st.rerun()

        except Exception as e: st.error(f"錯誤: {e}")

# === 10. 即時報表產生引擎 ===
st.divider()
st.subheader("📋 即時運算加工明細表")

display_data = []

for raw_idx, item in enumerate(st.session_state.get('raw_data_list', [])):
    b_len = item['base_len']
    count = item['count']
    lap = item['lap_len']
    uw = item['uw']
    
    if item['mode'] == 'stirrup' or b_len <= stock_len:
        display_data.append({
            "raw_idx": raw_idx, "番號": item['size_key'], "形狀": item['shape_str'],
            "單支長": round(b_len, 1), "支數": int(count), "總長(cm)": round(b_len * count, 1),
            "單位重": uw, "總重": round((b_len/100)*uw*count, 2), "備註": item['note']
        })
    else:
        if auto_split:
            pieces = split_rebar(b_len, stock_len, lap)
            for p_idx, p_len in enumerate(pieces):
                part_note = item['note'] + f" (Part {p_idx+1}/{len(pieces)} {'定尺' if p_len==stock_len else '餘料'})"
                display_data.append({
                    "raw_idx": raw_idx, "番號": item['size_key'], "形狀": item['shape_str'],
                    "單支長": round(p_len, 1), "支數": int(count), "總長(cm)": round(p_len * count, 1),
                    "單位重": uw, "總重": round((p_len/100)*uw*count, 2), "備註": part_note
                })
        else:
            laps = math.floor(b_len / stock_len)
            if b_len % stock_len == 0: laps -= 1
            total_merge_len = b_len + laps * lap
            merge_note = item['note'] + f" (含搭接{laps}處)"
            display_data.append({
                "raw_idx": raw_idx, "番號": item['size_key'], "形狀": item['shape_str'],
                "單支長": round(total_merge_len, 1), "支數": int(count), "總長(cm)": round(total_merge_len * count, 1),
                "單位重": uw, "總重": round((total_merge_len/100)*uw*count, 2), "備註": merge_note
            })

if display_data:
    df = pd.DataFrame(display_data)
    st.markdown("#### 📊 總量統計")
    summary = df.groupby("番號")["總重"].sum().reset_index()
    summary["噸數"] = summary["總重"] / 1000; summary["金額"] = summary["噸數"] * unit_price
    st.dataframe(summary.style.format({"總重": "{:.2f}", "噸數": "{:.3f}", "金額": "${:,.0f}"}), use_container_width=True)
    
    st.markdown("#### 📄 加工裁切明細")
    cols = st.columns([0.5, 1, 1.5, 1, 1, 1.5, 1, 1.5, 3, 0.5])
    headers = ["#","番號","形狀","單長(cm)","支數","總長(cm)","單位重","總重","備註",""]
    for c, h in zip(cols, headers): c.markdown(f"**{h}**")
    
    for i, row in df.iterrows():
        cols = st.columns([0.5, 1, 1.5, 1, 1, 1.5, 1, 1.5, 3, 0.5])
        cols[0].write(f"{i+1}"); cols[1].write(row['番號']); cols[2].write(row['形狀'])
        len_str = f"<span style='color:red; font-weight:bold'>{row['單支長']}</span>" if row['單支長'] > stock_len else f"{row['單支長']}"
        cols[3].markdown(len_str, unsafe_allow_html=True)
        cols[4].write(f"{row['支數']}"); cols[5].write(f"{row['總長(cm)']}")
        cols[6].write(f"{row['單位重']}"); cols[7].write(f"{row['總重']}"); cols[8].write(row['備註'])
        if cols[9].button("🗑️", key=f"del_{i}"): delete_item(row['raw_idx']); st.rerun()

    st.markdown("---")
    col_del, col_dl = st.columns([1, 4])
    with col_del:
        if st.button("🗑️ 清空全部", type="secondary"): clear_all_data(); st.rerun()
    with col_dl:
        def export_excel():
            wb = Workbook(); ws = wb.active; ws.title = "撿料單"
            font_header = Font(name='微軟正黑體', bold=True, size=12)
            font_body = Font(name='微軟正黑體', size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.merge_cells('A1:D1'); ws['A1'] = f"建案名稱: {project_name}"; ws['A1'].font = font_header
            ws.merge_cells('E1:H1'); ws['E1'] = f"聯絡人: {contact_person}"; ws['E1'].font = font_header
            ws.merge_cells('I1:L1'); ws['I1'] = f"結構部位: {structure_part}"; ws['I1'].font = font_header
            headers = ["編號","番號","形狀","單支長\n(cm)","支數","總長\n(cm)","單位重","總重\n(kg)","備註"]
            for i, t in enumerate(headers, 1):
                c = ws.cell(row=2, column=i, value=t); c.font = font_header; c.border = border; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for r, row in df.iterrows():
                d = [r+1, row['番號'], row['形狀'], row['單支長'], row['支數'], row['總長(cm)'], row['單位重'], row['總重'], row['備註']]
                for i, v in enumerate(d, 1): c = ws.cell(row=r+3, column=i, value=v); c.font = font_body; c.border = border; c.alignment = Alignment(horizontal='center', vertical='center')
            return wb
        out = BytesIO(); wb = export_excel(); wb.save(out)
        st.download_button("📥 下載加工 Excel", out.getvalue(), f"{project_name}_下料單.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

else: st.info("目前無資料，請從上方新增。")
